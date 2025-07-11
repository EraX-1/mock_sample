import argparse
import json
import os

from InquirerPy import prompt
from openpyxl import Workbook, load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from tqdm import tqdm

from src.evaluation.methods import run_evaluation
from src.internal.searcher import semantic_hybrid_search

TMP_BASE_FILE_PATH = "tmp/evaluation"


def semantic_hybrid_search_mock(query, index_type):
    return semantic_hybrid_search(
        query=query,
        index_type=index_type,
        custom_prompt="",
        is_active_custom_prompt=False,
        model="gpt-4o-mini",
        history=[],
        from_job=True,
    )


def create_evaluation_directory():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(current_dir)
    tmp_dir = os.path.join(parent_dir, "tmp")
    evaluation_dir = os.path.join(tmp_dir, "evaluation")
    os.makedirs(evaluation_dir, exist_ok=True)
    input_dir = os.path.join(evaluation_dir, "input")
    output_dir = os.path.join(evaluation_dir, "output")
    command_dir = os.path.join(evaluation_dir, "command")
    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(command_dir, exist_ok=True)


def prompt_user_input():
    questions = [
        {
            "type": "list",
            "name": "reference",
            "message": "参照元を選択してください:",
            "choices": ["NKJG", "JES", "GUIDE_STEP", "CHECK_LIST", "TASK_LIST"],
            "default": "NKJG",
        },
        {
            "type": "input",
            "name": "prompt",
            "message": "プロンプトを入力してください:",
        },
        {
            "type": "input",
            "name": "truth",
            "message": "期待する回答を入力してください:",
        },
        {
            "type": "input",
            "name": "relevant",
            "message": "本来回答に使用したいデータを記述してください。:",
        },
    ]
    answers = prompt(questions)
    return (
        answers["reference"],
        answers["prompt"],
        answers["truth"],
        answers["relevant"],
    )


def select_file():
    file_list = os.listdir(f"{TMP_BASE_FILE_PATH}/input")
    questions = [
        {
            "type": "list",
            "name": "file",
            "message": "ファイルを選択してください:",
            "choices": [file for file in file_list],
        }
    ]
    answers = prompt(questions)
    return answers["file"]


def parse_arguments():
    parser = argparse.ArgumentParser(description="チャット回答を評価するコマンド")
    parser.add_argument("--reference", type=str, help="参照元を指定 (例: NKJG)")
    parser.add_argument("--prompt", type=str, help="プロンプトを指定")
    parser.add_argument("--truth", type=str, help="期待する回答を指定")
    parser.add_argument("--relevant", type=str, help="本来使用したいデータを指定")
    parser.add_argument("--excel", action="store_true", help="Excelファイルインポートを使用")
    parser.add_argument(
        "--file",
        type=str,
        help="Excelファイルインポートを使用する場合、パスを指定（/appからの相対パス）",
    )
    parser.add_argument("--output", action="store_true", help="Excelファイルエクスポートを使用")
    args = parser.parse_args()
    return args


# evaluation/example.py から移植
def extract_content(json_data: dict) -> list[str]:
    return [item["content"] for item in json_data["queryResult"]["value"]]


def extract_docs_from_file(
    file_path: str,
) -> tuple[list[str], list[str], str, str, str, str]:
    with open(file_path, encoding="utf-8") as f:
        result = json.load(f)
    return (
        result["relevantDocuments"],
        extract_content(result),
        result["query"],
        result["input_prompt"],
        result["truth"],
        result["generated"],
    )


def load_evaluation_data(
    file_paths: list[str],
) -> tuple[list[str], dict[str, list[str]], dict[str], dict[str], dict[str]]:
    queries = []
    retriever_results = {}
    ground_truth_relevant_docs = {}

    for file_path in file_paths:
        (
            relevant_docs,
            retrieved_docs,
            query,
            input_prompt,
            truth,
            generated,
        ) = extract_docs_from_file(file_path)
        queries.append(query)
        retriever_results[query] = retrieved_docs
        ground_truth_relevant_docs[query] = relevant_docs
        input_prompt = input_prompt
        truth = truth
        generated_response = generated

    return (
        queries,
        retriever_results,
        ground_truth_relevant_docs,
        input_prompt,
        truth,
        generated_response,
    )


def normalize_text(text: str):
    return (
        text.replace("\n", "")
        .replace(" ", "")
        .replace("　", "")
        .replace("\\n", "")
        .replace("\r", "")
    )


if __name__ == "__main__":
    create_evaluation_directory()
    args = parse_arguments()
    # xlsxファイルからの一括評価
    if args.excel:
        if args.file:
            filename = args.file
        else:
            filename = select_file()
            filename = f"{TMP_BASE_FILE_PATH}/input/{filename}"
        workbook = load_workbook(filename)
        sheet = workbook.active
        evaluation_target_list = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            reference = row[0].value if row[0].value is not None else None
            query = row[1].value if row[1].value is not None else None
            truth = row[2].value if row[2].value is not None else None
            relevant = [cell.value for cell in row[3:] if cell.value is not None]
            # 3列すべてがデータ無しの場合、終了
            if reference is None and query is None and not relevant:
                break
            evaluation_target_list.append(
                {
                    "reference": reference,
                    "input_prompt": query,
                    "truth": truth,
                    "relevant_docs": relevant,
                }
            )
    # 引数で単発の情報が与えられている場合
    elif args.reference and args.prompt and args.truth:
        evaluation_target_list = [
            {
                "reference": args.reference,
                "input_prompt": args.prompt,
                "truth": args.truth,
                "relevant_docs": args.relevant,
            }
        ]

    # それ以外（対話形式で単発の評価）
    else:
        reference, input_prompt, truth, relevant = prompt_user_input()
        evaluation_target_list = [
            {
                "reference": reference,
                "input_prompt": input_prompt,
                "truth": truth,
                "relevant_docs": relevant,
            }
        ]
    print("推論を開始しました。しばらくお待ち下さい...")
    for index, evaluation_target in enumerate(tqdm(evaluation_target_list)):
        inference_data = semantic_hybrid_search_mock(
            query=evaluation_target["input_prompt"],
            index_type=evaluation_target["reference"],
        )
        evaluation_dict = {
            "queryResult": {
                "value": [
                    {
                        "id": doc["id"],
                        "keywords": None,
                        "content": normalize_text(doc["content"]),
                        "sourceFileName": doc["sourceFileName"],
                        "pageNumber": doc["pageNumber"],
                    }
                    for doc in inference_data["reference_docs"]
                ]
            },
            "input_prompt": evaluation_target["input_prompt"],
            "generated": inference_data["generated"],
            "truth": evaluation_target["truth"],
            "relevantDocuments": [
                normalize_text(relevant_doc)
                for relevant_doc in evaluation_target["relevant_docs"]
            ],
            "query": inference_data["query"],
        }
        with open(
            f"{TMP_BASE_FILE_PATH}/command/generated_{index}.json",
            "w",
            encoding="utf-8",
        ) as file:
            json.dump(evaluation_dict, file, ensure_ascii=False, indent=4)
    print("評価用ファイルを作成しました")
    file_paths = [
        f"{TMP_BASE_FILE_PATH}/command/generated_{index}.json"
        for index in range(len(evaluation_target_list))
    ]
    wb = Workbook()
    ws = wb.active
    ws.append(["期待回答", "実際回答", "コサイン類似度"])
    for file_path in file_paths:
        (
            queries,
            retriever_results,
            ground_truth_relevant_docs,
            input_prompt,
            truth,
            generated_responses,
        ) = load_evaluation_data([file_path])
        print("プロンプト:", input_prompt)
        print("===")
        print("期待する回答：", truth)
        print("===")
        print("生成された回答：", generated_responses)
        if ground_truth_relevant_docs:
            run_evaluation(queries, retriever_results, ground_truth_relevant_docs)
        vectorizer = TfidfVectorizer()
        tfidf_matrix = vectorizer.fit_transform([truth, generated_responses])

        # コサイン類似度を計算
        cosine_sim = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])
        ws.append([truth, generated_responses, cosine_sim[0][0]])
        print(f"コサイン類似度: {cosine_sim[0][0]:.4f}")
        print("************************")
    wb.save(f"{TMP_BASE_FILE_PATH}/output/evaluation.xlsx")
